# Written by Eric Martin for COMP9021


# Extracts from the list of countries accessible at
# http://www.worldbank.org/en/where-we-work
# the value of IBRD/IDA operations, when available,
# and creates a spreadsheet with those data.


import os.path
import sys
import urllib.request

import bs4
import openpyxl


def convert_to_number(amount):
    units = {'thousand': 10 ** 3, 'million': 10 ** 6,
             'billion': 10 ** 9, 'trillion': 10 ** 12
            }
    # amount is of the form ""$v " followed by one of the units above,
    # possibly preceded or followed by spaces, with v a floating point
    # value.
    for unit in units:
        if unit in amount:
            return int(float(amount.strip().lstrip('$').rstrip(unit))
                       * units[unit]
                      )

def countries_and_data():
    # Example of html code being matched:
    #
    # <a class="alpha-name dropdown-item"
    #    href="https://www.worldbank.org/en/country/afghanistan "
    # >Afghanistan</a>
    for country in top_page.select('a[class="alpha-name dropdown-item"]'):
        country_name = country.text
        try:
            with urllib.request.urlopen(country.get('href')) as overview_url:
                overview_page = bs4.BeautifulSoup(overview_url, 'html.parser')
                # Examples of html code being matched:
                #
                # <a href="https://projects.worldbank.org/search?
                #                   lang=en&searchTerm=&countrycode_exact=AF
                #         " class="lp__card_link">Projects
                # </a>
                #
                # <a href="https://data.worldbank.org/country/afghanistan"
                #    class="lp__card_link">Data
                # </a>
                # 
                for data in overview_page.select('a[class="lp__card_link"]'):
                    # On some pages, we find DATA instead of Data.
                    if data.string.lower() == 'data':
                        break
                else:
                    print(f'No Data page for {country_name}.')
                    continue
                try:
                    with urllib.request.urlopen(data.get('href')) as data_url:
                        data_page = bs4.BeautifulSoup(data_url, 'html.parser')
                        # Example of html code being matched:
                        #
                        # <span class="name" data-reactid="293">
                        #       IBRD/IDA Operations Approved by Fiscal Year
                        # </span>
                        # <div class="chart" data-reactid="294">
                        #    <div class="chart-summry" data-reactid="295">
                        #       <div data-reactid="296">
                        #          <em data-reactid="297">$25.00 million</em>
                        #          ...
                        #       </div>
                        #    </div>
                        # </div>
                        #
                        # All continuous text, without embedded new lines,
                        # which would otherwise be children.
                        try:
                            yield country_name,\
                                  convert_to_number(next(next(next(
                                        data_page.find('span',
                                                       text=indicator
                                                      ).next_sibling.children
                                                                  ).children
                                                              ).children
                                                        ).string
                                                    )
                        except AttributeError:
                            print(f'No {indicator} for {country_name}.')
                except (urllib.error.HTTPError, urllib.error.URLError):
                    print('Could not access Data page for '
                          f'{country_name}.'
                         )
        except (urllib.error.HTTPError, urllib.error.URLError):
            print(f'Could not access overview page for {country_name}.')

spreadsheet_file_name = 'IBRD_IDA_operations.xlsx'
indicator = 'IBRD/IDA Operations Approved by Fiscal Year'
if os.path.isfile(spreadsheet_file_name):
    print(f'A file named {spreadsheet_file_name} already exists.') 
    print('You have to remove it first.')
    sys.exit()
try:
    with urllib.request.urlopen('http://www.worldbank.org/en/where-we-work')\
            as top_url:
        top_page = bs4.BeautifulSoup(top_url, 'html.parser')
        workbook = openpyxl.Workbook()
        spreadsheet = workbook.active
        spreadsheet.title = 'World countries'
        spreadsheet['A1'] = 'Country'
        spreadsheet['B1'] = 'IBRD/IDA operations'
        for counter, (country, amount) in enumerate(countries_and_data(), 2):
            spreadsheet.cell(row=counter, column=1).value = country
            spreadsheet.cell(row=counter, column=2).value = amount
        workbook.save(spreadsheet_file_name)
except (urllib.error.HTTPError, urllib.error.URLError):
    print('Could not access the top resource.')
